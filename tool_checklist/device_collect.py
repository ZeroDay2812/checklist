import argparse
import os
import re
import requests
import json
import urllib3
import openpyxl

from ansible import context
from ansible.cli import CLI
from ansible.module_utils.common.collections import ImmutableDict
from ansible.executor.playbook_executor import PlaybookExecutor
from ansible.parsing.dataloader import DataLoader
from ansible.inventory.manager import InventoryManager
from ansible.vars.manager import VariableManager
from requests.api import patch
from pprint import pprint
from openpyxl.styles.borders import Border, Side

CURRENT_DIR = os.getcwd()


def call_api_get(url, headers):
    try:
        response = requests.get(url, headers=headers, verify=False)
    except Exception as e:
        print(e)
        return False
    if str(response.status_code).startswith('4'):
        print("[Fail] Can not get information at url: {}. "
              "Code {}".format(url, response.status_code))
    return response.json()


def hpe_get_session_token(base_url, username, password):
    headers = {"Content-Type": "application/json", "OData-Version": "4.0"}
    rq_data = json.dumps({"UserName": username, "Password": password})
    url = base_url + "/redfish/v1/SessionService/Sessions/"
    try:
        response = requests.post(url, headers=headers, data=rq_data,
                                 verify=False)
    except Exception as e:
        print(e)
        return False
    if response.status_code != 201:
        print("[Fail] Can not get session token at url: {}. "
              "Code {}".format(url, response.status_code))
        return False
    print("[Success] Created session at url: {}".format(url))
    return response.headers


def hpe_expire_session_token(url, headers):
    try:
        response = requests.delete(url, headers=headers, verify=False)
    except Exception as e:
        print(e)
        return False
    if response.status_code != 200:
        print("[Fail] Can not expire session at url: {}. "
              "Code {}".format(url, response.status_code))
    print("[Success] Expired session at url: {}".format(url))
    return None


def hpe_get_processor_info(base_url, headers):
    try:
        processors = call_api_get(
            base_url + "/redfish/v1/Systems/1/Processors/", headers)
        if not processors:
            return False
        result, info, fail_part = {}, [], []
        for i in processors.get('Members'):
            cpu_url = base_url + i['@odata.id']
            cpu = call_api_get(cpu_url, headers)
            if not cpu:
                continue
            if cpu['Status']['Health'] != 'OK':
                fail_part.append({
                    'location': cpu.get('Socket'),
                    'health': cpu['Status']['Health'],
                    'state': cpu['Status']['State']
                })
            cpu_data = {
                'core': cpu.get('TotalCores'),
                'thread': cpu.get('TotalThreads'),
                'model': cpu.get('Model')
            }
            info.append(cpu_data)
        result.update({'fail_part': fail_part, 'info': info,
                       'number': len(processors.get('Members'))})
        return result
    except Exception as e:
        print("Can not get processor info cause {}".format(e))
        return None


def hpe_get_fan_info(base_url, headers):
    try:
        thermal = call_api_get(base_url + "/redfish/v1/Chassis/1/Thermal/",
                               headers)
        if not thermal:
            return False
        result = {}
        result.update({'number': len(thermal.get('Fans'))})
        fail_part = []
        for i in thermal.get('Fans'):
            if i['Status'].get('State') != 'Absent':
                if i['Status'].get('Health') != 'OK':
                    fail_part.append({
                        'name': i.get('Name'),
                        'health': i['Status'].get('Health'),
                        'state': i['Status'].get('State')
                    })
        result.update({'fail_part': fail_part})
        return result
    except Exception as e:
        print("Can not get fan info cause {}".format(e))
        return None


def hpe_get_memory_info(base_url, headers):
    try:
        mem = call_api_get(base_url + "/redfish/v1/Systems/1/Memory/", headers)
        if not mem:
            return False
        result, info, fail_part = {}, [], []
        for i in mem.get('Members'):
            ram = call_api_get(base_url + i['@odata.id'], headers)
            if not ram:
                continue
            if ram.get('Status'):
                if ram['Status']['State'] == 'Absent':
                    continue
                if ram['Status']['Health'] != 'OK':
                    fail_part.append({
                        'location': ram.get('DeviceLocator'),
                        'health': ram['Status']['Health'],
                        'state': ram['Status']['State']
                    })
            capa = ram.get('CapacityMiB')
            if not capa:
                capa = ram.get('SizeMB')
            type = ram.get('MemoryDeviceType')
            if not type:
                type = ram.get('DIMMType')
            if not info:
                info.append({
                    'index': 0,
                    'capacity_mib': capa,
                    'type': type,
                    'count': 1
                })
            else:
                new = True
                for r in info:
                    count = r['count']
                    index = r['index']
                    if r['capacity_mib'] == capa:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                if new:
                    info.append({
                        'index': len(info),
                        'capacity_mib': capa,
                        'type': type,
                        'count': 1
                    })
        result.update({'fail_part': fail_part, 'info': info})
        return result
    except Exception as e:
        print("Can not get memory info cause {}".format(e))
        return None


def hpe_get_logical_disk_info(base_url, storage, headers):
    try:
        all_ldisk = call_api_get(base_url + storage + "LogicalDrives/",
                                 headers)
        if not all_ldisk:
            return False
        result = {}
        result.update({'number': len(all_ldisk.get('Members'))})
        info, ldisk_data = [], {}
        for i in all_ldisk.get('Members'):
            ldisk_url = base_url + i['@odata.id']
            ldisk = call_api_get(ldisk_url, headers)
            if not ldisk:
                continue
            ldisk_data = {
                'raid': ldisk.get('Raid'),
                'capacity_gb': int(round(ldisk['CapacityMiB'] / 954)),
            }
            pdisk_url = ldisk_url + 'DataDrives/'
            pdisk = call_api_get(pdisk_url, headers)
            if not pdisk:
                ldisk_data.update({'disks': []})
                info.append(ldisk_data)
                continue
            disk = []
            ldisk_data.update({'physical_count': len(pdisk['Members'])})
            for d in pdisk.get('Members'):
                pdisk_info = call_api_get(base_url + d['@odata.id'], headers)
                if not pdisk_info:
                    continue
                disk_data = {
                    'capacity_gb': pdisk_info['CapacityGB'],
                    'type': pdisk_info['MediaType']
                }
                if disk_data['type'] == 'HDD':
                    disk_data.update({
                        'speed': pdisk_info['RotationalSpeedRpm']})
                else:
                    disk_data.update({'speed': None})
                if not disk:
                    disk.append(disk_data)
                for r in disk:
                    if disk_data['capacity_gb'] != r['capacity_gb'] \
                            or disk_data['type'] != r['type']:
                        disk.append(disk_data)
            ldisk_data.update({'disks': disk})
            info.append(ldisk_data)
        result.update({'info': info})
        return result
    except Exception as e:
        print("Can not get logical disk info cause {}".format(e))
        return None


def hpe_get_physical_disk_info(base_url, storage, headers):
    try:
        all_pdisk = call_api_get(base_url + storage + "DiskDrives/", headers)
        if not all_pdisk:
            return False
        result = {}
        result.update({'number': len(all_pdisk.get('Members'))})
        info, fail_part = [], []
        for i in all_pdisk.get('Members'):
            pdisk = call_api_get(base_url + i['@odata.id'], headers)
            if not pdisk:
                continue
            if pdisk['Status']['Health'] != 'OK' \
                    or pdisk['Status']['State'] != 'Enabled':
                location = pdisk['Location'].split(":")
                location_format = pdisk['LocationFormat'].split(":")
                lmsg = ''
                for lo in range(len(location)):
                    lmsg = location_format[lo] + ' ' + location[lo] + ' '
                fail_part.append(lmsg.strip())
            disk_data = {
                'capacity_gb': pdisk['CapacityGB'],
                'type': pdisk['MediaType']
            }
            if disk_data['type'] == 'HDD':
                disk_data.update({'speed': pdisk['RotationalSpeedRpm']})
            else:
                disk_data.update({'speed': None})
            if not info:
                disk_data.update({'count': 1, 'index': 0})
                info.append(disk_data)
            else:
                new = True
                for d in info:
                    count = d['count']
                    index = d['index']
                    if d['speed'] \
                            and d['capacity_gb'] == disk_data['capacity_gb'] \
                            and d['type'] == disk_data['type'] \
                            and d['speed'] == disk_data['speed']:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                    elif not d['speed'] \
                            and d['capacity_gb'] == disk_data['capacity_gb'] \
                            and d['type'] == disk_data['type']:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                if new:
                    disk_data.update({'count': 1, 'index': len(info)})
                    info.append(disk_data)
        result.update({'fail_part': fail_part, 'info': info})
        return result
    except Exception as e:
        print("Can not get physical disk info cause {}".format(e))
        return None


def hpe_get_disk_info(base_url, headers):
    try:
        raid_controller_info = call_api_get(
            base_url + "/redfish/v1/Systems/1/SmartStorage/ArrayControllers/",
            headers)
        if not raid_controller_info:
            return False
        if raid_controller_info.get('Members') \
                and len(raid_controller_info.get('Members')) == 1:
            storage = raid_controller_info.get('Members')[0]['@odata.id']
        else:
            return False
        ldisk = hpe_get_logical_disk_info(base_url, storage, headers)
        pdisk = hpe_get_physical_disk_info(base_url, storage, headers)
        return {'logical_disk': ldisk, 'physical_disk': pdisk}
    except Exception as e:
        print("Can not get disk info cause {}".format(e))
        return None


def hpe_get_power_info(base_url, headers):
    all_power = call_api_get(base_url + "/redfish/v1/Chassis/1/Power/",
                             headers)
    if not all_power or not all_power.get('PowerSupplies'):
        return False
    result, info, fail_part = {}, [], []
    try:
        for i in all_power.get('PowerSupplies'):
            if i['Status']['Health'] != 'OK' \
                    or i['Status']['State'] != 'Enabled':
                fp = {'id': i.get('MemberId'),
                      'health': i['Status']['Health'],
                      'state': i['Status']['State']}
                fail_part.append(fp)
            power_data = {
                'capacity': i['PowerCapacityWatts'],
            }
            info.append(power_data)
        redundancy = False
        if all_power.get('Redundancy'):
            for re in all_power.get('Redundancy'):
                if re.get('Mode') and re.get('Mode') == 'Failover':
                    redundancy = True
        result.update({'number': len(all_power['PowerSupplies']), 'info': info,
                       'fail_part': fail_part, 'redundancy': redundancy})
        return result
    except Exception as e:
        print('Can not get Power info cause {}'.format(e))
        return None


def hpe_get_network_adapter_info(base_url, headers, ilo):
    try:
        if not ilo or ilo != 'iLO 4':
            all_na = call_api_get(
                base_url + "/redfish/v1/Systems/1/BaseNetworkAdapters/",
                headers)
        else:
            all_na = call_api_get(
                base_url + "/redfish/v1/Systems/1/NetworkAdapters/",
                headers)
        if not all_na:
            return False
        result = {}
        result.update({'number': len(all_na.get('Members'))})
        info, fail_part = [], []
        for i in all_na.get('Members'):
            na_url = base_url + i['@odata.id']
            na = call_api_get(na_url, headers)
            if not na:
                continue
            na_data = {
                'model': na.get('Name')
            }
            if not na.get('Status'):
                info.append(na_data)
                continue
            if na['Status'].get('Health') \
                    and na['Status'].get('Health') != 'OK':
                fail_part.append({
                    'model': na.get('Name'), 'health': na['Status']['Health'],
                    'state': na['Status'].get('State')
                })
            info.append(na_data)
        result.update({'info': info, 'fail_part': fail_part})
        return result
    except Exception as e:
        print("Can not get Network info cause {}".format(e))
        return None


def hpe_get_ilo_info(base_url, headers):
    try:
        manager_info = call_api_get(base_url + "/redfish/v1/Managers/1/",
                                    headers)
        if not manager_info:
            return False
        firm_version = manager_info.get('FirmwareVersion')
        if firm_version.startswith('iLO 4'):
            firm = manager_info['Oem']['Hp']['Firmware']['Current']
            ilo = 'iLO 4'
            firm_string = firm['VersionString'].replace(ilo, "").strip() \
                          + ' ' + firm['Date']
        else:
            firm = manager_info['Oem']['Hpe']['Firmware']['Current']
            ilo = 'iLO ' + firm['VersionString'].split()[1]
            firm_string = firm['VersionString'].replace(ilo, "").strip() \
                          + ' ' + firm['Date']
        result = {'model': ilo, 'version': firm_string}
        return result
    except Exception as e:
        print("Can not get iLO version cause {}".format(e))
        return None


def hpe_get_snmp_service_info(base_url, headers):
    try:
        snmp_info = call_api_get(
            base_url + "/redfish/v1/Managers/1/SnmpService/", headers)
        if not snmp_info:
            return False
        result = {
            'state': snmp_info['Status']['State'],
            'list_string': snmp_info.get('ReadCommunities')
        }
        return result
    except Exception as e:
        print("Can not get SNMP info cause {}".format(e))
        return None


def hpe_get_bios_config_info(base_url, headers):
    try:
        bios = call_api_get(base_url + "/redfish/v1/Systems/1/Bios/", headers)
        if not bios:
            return False
        attribute = bios.get('Attributes')
        if not attribute:
            result = {
                'thermal': bios.get('ThermalConfig'),
                'workload_profile': bios.get('PowerProfile'),
                'power_performance': bios.get('PowerRegulator')
            }
        else:
            result = {
                'thermal': attribute.get('ThermalConfig'),
                'power_performance': attribute.get('PowerRegulator')
            }
        if result['thermal'] == 'OptimalCooling' and result[
            'power_performance'] == 'StaticHighPerf':
            return "OK. " + str(result)
        else:
            return "NOK. " + str(result)
    except Exception as e:
        print("Can not get BIOS config info cause {}".format(e))
        return None


def hpe_get_basic_info(base_url, headers):
    try:
        basic_info = call_api_get(base_url + "/redfish/v1/Systems/1/", headers)
        if not basic_info:
            return False
        memory_status = basic_info['MemorySummary']['Status'].get(
            'HealthRollUp')
        if not memory_status:
            memory_status = basic_info['MemorySummary']['Status'].get(
                'HealthRollup')
        result = {
            'bios_version': basic_info.get('BiosVersion'),
            'model': basic_info.get('Model'),
            'memory_gib': basic_info['MemorySummary']['TotalSystemMemoryGiB'],
            'memory_status': memory_status,
            'processor': {
                'number': basic_info['ProcessorSummary']['Count'],
                'model': basic_info['ProcessorSummary']['Model']
            },
            'serial': basic_info.get('SerialNumber'),
            'health': basic_info['Status']['Health']
        }
        return result
    except Exception as e:
        print("Can not get basic info cause {}".format(e))
        return None


def hpe_get_all_info(ip_address, username, password):
    print("--> Getting information for HPE device {}".format(ip_address))
    base_url = "https://" + ip_address
    header_info = hpe_get_session_token(base_url, username, password)
    if not header_info:
        return None
    token = header_info.get("X-Auth-Token")
    headers = {"Content-Type": "application/json", "X-Auth-Token": token}
    try:
        ilo = hpe_get_ilo_info(base_url, headers)
        if ilo:
            network = hpe_get_network_adapter_info(
                base_url, headers, ilo.get('model'))
        else:
            network = None
        result = {
            'ip': ip_address,
            'base_info': hpe_get_basic_info(base_url, headers),
            'firmware': ilo,
            'network': network,
            'processor': hpe_get_processor_info(base_url, headers),
            'fan': hpe_get_fan_info(base_url, headers),
            'disk': hpe_get_disk_info(base_url, headers),
            'power': hpe_get_power_info(base_url, headers),
            'memory': hpe_get_memory_info(base_url, headers),
            'snmp': hpe_get_snmp_service_info(base_url, headers),
            'bios_config': hpe_get_bios_config_info(base_url, headers)
        }
    except Exception as e:
        pprint(e)
        result = None

    hpe_expire_session_token(header_info.get('Location'), headers)
    return result


def fjs_get_session_token(base_url, username, password):
    headers = {"Content-Type": "application/json", "OData-Version": "4.0"}
    auth = (username, password)
    rq_data = json.dumps({"UserName": username, "Password": password})
    url = base_url + "/redfish/v1/SessionService/Sessions/"
    try:
        response = requests.post(url, headers=headers, data=rq_data,
                                 auth=auth, verify=False)
    except Exception as e:
        print(e)
        return False
    if response.status_code != 201:
        print("[Fail] Can not get session token at url: {}. "
              "Code {}".format(url, response.status_code))
        return False
    print("[Success] Created session at url: {}".format(url))
    return response.headers


def fjs_expire_session_token(url, headers):
    try:
        response = requests.delete(url, headers=headers, verify=False)
    except Exception as e:
        print(e)
        return False
    if response.status_code != 200:
        print("[Fail] Can not expire session at url: {}. "
              "Code {}".format(url, response.status_code))
    print("[Success] Expired session at url: {}".format(url))
    return None


def fjs_get_basic_info(system_url, headers):
    basic_info = call_api_get(system_url, headers)
    if not basic_info:
        return False
    try:
        memory_status = basic_info['MemorySummary']['Status'].get('Health')
        if not memory_status:
            memory_status = basic_info['MemorySummary']['Status'].get(
                'HealthRollup')
        result = {
            'bios_version': basic_info.get('BiosVersion'),
            'model': basic_info.get('Model'),
            'memory_gib': basic_info['MemorySummary']['TotalSystemMemoryGiB'],
            'memory_status': memory_status,
            'processor': {
                'number': basic_info['ProcessorSummary']['Count'],
                'model': basic_info['ProcessorSummary']['Model']
            },
            'serial': basic_info.get('SerialNumber'),
            'health': basic_info['Status']['Health']
        }
        return result
    except Exception as e:
        print("Can not get basic info because {}".format(e))
        return False


def fjs_get_processor_info(base_url, system, headers):
    processor = call_api_get(base_url + system + "/Processors/", headers)
    if not processor:
        return False
    try:
        result, info, fail_part = {}, [], []
        for i in processor.get('Members'):
            cpu_url = base_url + i['@odata.id']
            cpu = call_api_get(cpu_url, headers)
            if not cpu:
                continue
            if cpu['Status']['Health'] != 'OK':
                fail_part.append({
                    'location': cpu.get('Socket'),
                    'health': cpu['Status']['Health'],
                    'state': cpu['Status']['State']
                })
            cpu_data = {
                'core': cpu.get('TotalCores'),
                'thread': cpu.get('TotalThreads'),
                'model': cpu.get('Model')
            }
            info.append(cpu_data)
        result.update({'fail_part': fail_part, 'info': info,
                       'number': len(processor.get('Members'))})
        return result
    except Exception as e:
        print("Can not get Processor info because {}".format(e))
        return False


def fjs_get_fan_info(chassis_url, headers):
    thermal = call_api_get(chassis_url + "/Thermal/", headers)
    if not thermal:
        return False
    try:
        result, fail_part, number = {}, [], 0
        for i in thermal.get('Fans'):
            if i['Status'].get('State') != 'Absent':
                number += 1
                if i['Status'].get('Health') != 'OK':
                    fail_part.append({
                        'name': i.get('Name'),
                        'health': i['Status'].get('Health'),
                        'state': i['Status'].get('State')
                    })
        result.update({'number': number, 'fail_part': fail_part})
        return result
    except Exception as e:
        print("Can not get Fan info because {}".format(e))
        return False


def fjs_get_memory_info(base_url, system, headers):
    mem = call_api_get(base_url + system + "/Memory/", headers)
    if not mem:
        return False
    result, info, fail_part = {}, [], []
    try:
        for i in mem.get('Members'):
            ram = call_api_get(base_url + i['@odata.id'], headers)
            if not ram:
                continue
            if ram.get('Status'):
                if ram['Status']['State'] == 'Absent':
                    continue
                if ram['Status']['Health'] != 'OK':
                    fail_part.append({
                        'location': ram.get('DeviceLocator'),
                        'health': ram['Status']['Health'],
                        'state': ram['Status']['State']
                    })
            capa = ram.get('CapacityMiB')
            if not capa:
                capa = ram.get('SizeMB')
            type = ram.get('MemoryDeviceType')
            if not type:
                type = ram.get('DIMMType')
            if not info:
                info.append({
                    'index': 0,
                    'capacity_mib': capa,
                    'type': type,
                    'count': 1
                })
            else:
                new = True
                for r in info:
                    count = r['count']
                    index = r['index']
                    if r['capacity_mib'] == capa:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                if new:
                    info.append({
                        'index': len(info),
                        'capacity_mib': capa,
                        'type': type,
                        'count': 1
                    })
        result.update({'fail_part': fail_part, 'info': info})
        return result
    except Exception as e:
        print("Can not get Memory info because {}".format(e))
        return False


def fjs_get_logical_disk_info(base_url, volume_url, headers):
    all_ldisk = call_api_get(base_url + volume_url, headers)
    if not all_ldisk:
        return False
    result = {}
    result.update({'number': len(all_ldisk.get('Members'))})
    info, ldisk_data = [], {}
    try:
        for i in all_ldisk.get('Members'):
            ldisk_url = base_url + i['@odata.id']
            ldisk = call_api_get(ldisk_url, headers)
            if not ldisk:
                continue
            ldisk_data = {
                'raid': ldisk.get('RAIDType').replace("RAID", ""),
                'capacity_gb': int(round(ldisk['CapacityBytes'] / 1000000000)),
            }
            pdisk = ldisk.get('Links')
            if pdisk:
                pdisk = pdisk.get('Drives')
            else:
                ldisk_data.update({'disks': []})
                info.append(ldisk_data)
                continue
            disk = []
            ldisk_data.update({'physical_count': len(pdisk)})
            for d in pdisk:
                pdisk_info = call_api_get(base_url + d['@odata.id'], headers)
                if not pdisk_info:
                    continue
                disk_data = {
                    'capacity_gb': int(
                        pdisk_info['CapacityBytes'] / 1000000000),
                    'type': pdisk_info['MediaType']
                }
                if disk_data['type'] == 'HDD':
                    disk_data.update({'speed': pdisk_info['RotationSpeedRPM']})
                else:
                    disk_data.update({'speed': None})
                if not disk:
                    disk.append(disk_data)
                for r in disk:
                    if disk_data['capacity_gb'] != r['capacity_gb'] \
                            or disk_data['type'] != r['type']:
                        disk.append(disk_data)
            ldisk_data.update({'disks': disk})
            info.append(ldisk_data)
        result.update({'info': info})
        return result
    except Exception as e:
        print("Can not get logical disk info because {}".format(e))
        return False


def fjs_get_physical_disk_info(base_url, drives, headers):
    result = {}
    result.update({'number': len(drives)})
    info, fail_part = [], []
    try:
        for i in drives:
            pdisk = call_api_get(base_url + i['@odata.id'], headers)
            if not pdisk:
                continue
            if pdisk['Status']['Health'] != 'OK' \
                    or pdisk['Status']['State'] != 'Enabled':
                location = pdisk['Location']['Info'] \
                    .replace("[", "").replace("]", "").split(":")
                location_format = pdisk['Location']['InfoFormat'] \
                    .replace("[", "").replace("]", "").split(":")
                lmsg = ''
                for lo in range(len(location)):
                    lmsg = strip(location_format[lo]) + ' ' \
                           + srtip(location[lo]) + ' '
                fail_part.append(strip(lmsg))
            disk_data = {
                'capacity_gb': int(pdisk['CapacityBytes'] / 1000000000),
                'type': pdisk['MediaType']
            }
            if disk_data['type'] == 'HDD':
                disk_data.update({'speed': pdisk['RotationSpeedRPM']})
            else:
                disk_data.update({'speed': None})
            if not info:
                disk_data.update({'count': 1, 'index': 0})
                info.append(disk_data)
            else:
                new = True
                for d in info:
                    count = d['count']
                    index = d['index']
                    if d['speed'] \
                            and d['capacity_gb'] == disk_data['capacity_gb'] \
                            and d['type'] == disk_data['type'] \
                            and d['speed'] == disk_data['speed']:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                    elif not d['speed'] \
                            and d['capacity_gb'] == disk_data['capacity_gb'] \
                            and d['type'] == disk_data['type']:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                if new:
                    disk_data.update({'count': 1, 'index': len(info)})
                    info.append(disk_data)

        result.update({'fail_part': fail_part, 'info': info})
        return result
    except Exception as e:
        print("Can not get physical disk info because {}".format(e))
        return False


def fjs_get_disk_info(base_url, system, headers):
    raid_controller_info = call_api_get(base_url + system + "/Storage/",
                                        headers)
    if not raid_controller_info:
        return False
    if raid_controller_info.get('Members') \
            and len(raid_controller_info.get('Members')) == 1:
        storage_url = raid_controller_info.get('Members')[0]['@odata.id']
    else:
        print("Server has more than 1 raid controller, fail to get Disk info")
        return False
    storage = call_api_get(base_url + storage_url, headers)
    if not storage:
        return False
    if storage.get('Volumes'):
        ldisk = fjs_get_logical_disk_info(base_url,
                                          storage['Volumes'].get('@odata.id'),
                                          headers)
    else:
        ldisk = None
    if storage.get('Drives'):
        pdisk = fjs_get_physical_disk_info(base_url, storage['Drives'],
                                           headers)
    else:
        pdisk = None
    return {'logical_disk': ldisk, 'physical_disk': pdisk}


def fjs_get_power_info(base_url, chassis, headers):
    all_power = call_api_get(base_url + chassis + "/Power/", headers)
    if not all_power:
        return False
    result, info, fail_part = {}, [], []
    try:
        result.update({
            'number': len(all_power.get('PowerSupplies'))
        })
        for i in all_power.get('PowerSupplies'):
            if i['Status']['Health'] != 'OK' or \
                    i['Status']['State'] != 'Enabled':
                fp = {'name': i.get('Name')}
                fail_part.append(fp)
            power_data = {
                'capacity': i['PowerCapacityWatts'],
            }
            info.append(power_data)
        redundancy = False
        if all_power.get('PowerControl'):
            redundancy = True
        result.update({'info': info, 'fail_part': fail_part,
                       'redundancy': redundancy})
        return result
    except Exception as e:
        print("Can not get Power info because {}".format(e))
        return False


def fjs_get_network_adapter_info(base_url, chassis, headers):
    all_na = call_api_get(base_url + chassis + "/NetworkAdapters/", headers)
    if not all_na:
        return False
    result, info, fail_part = {}, [], []
    try:
        for i in all_na.get('Members'):
            na_url = base_url + i['@odata.id']
            na = call_api_get(na_url, headers)
            if not na:
                continue
            na_data = {
                'model': na.get('Model'), 'name': na.get('Name')
            }
            if not na.get('Status'):
                info.append(na_data)
                continue
            if na['Status'].get('Health') \
                    and na['Status'].get('Health') != 'OK':
                fail_part.append({
                    'name': na.get('Name'), 'state': na['Status'].get('State'),
                    'health': na['Status']['Health'], 'model': na.get('Model')
                })
            info.append(na_data)
        result.update({'number': len(all_na.get('Members')), 'info': info,
                       'fail_part': fail_part})
        return result
    except Exception as e:
        print("Can not get Network info because {}".format(e))
        return False


def fjs_get_snmp_service_info(base_url, manager, headers):
    mn = call_api_get(base_url + manager + "/ManagerNetwork/", headers)
    if not mn:
        return False
    snmp = mn.get('SNMP')
    if not snmp:
        return False
    enable = snmp.get('ProtocolEnabled')
    if enable:
        result = {'state': 'Enabled', 'list_string': ''}
    else:
        result = {'state': 'Disabled', 'list_string': ''}
    return result


def fjs_get_irmc_info(base_url, manager, headers):
    irmc = call_api_get(base_url + manager, headers)
    if not irmc:
        return False
    result = {'model': irmc.get('Model'),
              'version': irmc.get('FirmwareVersion')}
    return result


def fjs_get_bios_config_info(base_url, headers):
    try:
        bios = call_api_get(base_url + "/rest/v1/Oem/eLCM/"
                                       "ProfileManagement/Server", headers)
        if not bios:
            return False
        cpuconfig = bios['Server']['SystemConfig']['BiosConfig']['CpuConfig']
        performance_mode = cpuconfig['EnergyPerformanceMode']
        os_energy = cpuconfig['OsEnergyPerformanceOverrideEnabled']
        if performance_mode == 'OptimizedPerformance' and os_energy == "True":
            return "OK. {'EnergyPerformanceMode': %s; " \
                   "'OsEnergyPerformanceOverrideEnabled': %s" \
                   % (performance_mode, os_energy)
        else:
            return "NOK. {'EnergyPerformanceMode': %s; " \
                   "'OsEnergyPerformanceOverrideEnabled': %s" \
                   % (performance_mode, os_energy)
    except Exception as e:
        print("Can not get BIOS config info cause {}".format(e))
        return None


def fjs_get_object_info(base_url, object, headers):
    object_dict = {
        'system': 'Systems',
        'chassis': 'Chassis',
        'manager': 'Managers'
    }
    oj = object_dict.get(object)
    if not oj:
        return False
    oj_info = call_api_get(base_url + "/redfish/v1/" + oj, headers)
    if not oj_info:
        return False
    oj_list = oj_info.get('Members')
    if not oj_list:
        return False
    if len(oj_list) != 1:
        print("Device doesnt have 1 {} only, can not get info".format(object))
        return False
    oj_url = oj_list[0].get('@odata.id')
    return oj_url


def fjs_get_all_info(ip_address, username, password):
    print("--> Getting information for Fujitsu device {}".format(ip_address))
    base_url = "https://" + ip_address
    header_info = fjs_get_session_token(base_url, username, password)
    if not header_info:
        return None
    token = header_info.get("X-Auth-Token")
    headers = {"Content-Type": "application/json", "X-Auth-Token": token}
    try:
        system = fjs_get_object_info(base_url, 'system', headers)
        chassis = fjs_get_object_info(base_url, 'chassis', headers)
        manager = fjs_get_object_info(base_url, 'manager', headers)
        result = {
            'ip': ip_address,
            'base_info': fjs_get_basic_info(base_url + system, headers),
            'network': fjs_get_network_adapter_info(
                base_url, chassis, headers),
            'processor': fjs_get_processor_info(base_url, system, headers),
            'fan': fjs_get_fan_info(base_url + chassis, headers),
            'disk': fjs_get_disk_info(base_url, system, headers),
            'power': fjs_get_power_info(base_url, chassis, headers),
            'memory': fjs_get_memory_info(base_url, system, headers),
            'snmp': fjs_get_snmp_service_info(base_url, manager, headers),
            'firmware': fjs_get_irmc_info(base_url, manager, headers),
            'bios_config': fjs_get_bios_config_info(base_url, headers)
        }
    except Exception as e:
        print("Can not get Fujitsu device info cause {}".format(e))
        result = None
    fjs_expire_session_token(base_url + header_info.get('Location'), headers)
    return result


def dell_get_redfish_version(base_url, username, password):
    auth = (username, password)
    try:
        response = requests.get(base_url + "/redfish/v1/", auth=auth,
                                verify=False)
        version = response.json().get('RedfishVersion')
        if version:
            version_num = int(version.replace(".", ""))
            return version_num
        return False
    except Exception as e:
        print("Can not get redfish version cause {}".format(e))
        return False


def dell_get_session_token(base_url, username, password, redfish_version):
    headers = {"Content-Type": "application/json"}
    auth = (username, password)
    rq_data = json.dumps({"UserName": username, "Password": password})
    if redfish_version >= 160:
        url = base_url + "/redfish/v1/SessionService/Sessions/"
    else:
        url = base_url + "/redfish/v1/Sessions/"
    try:
        response = requests.post(url, headers=headers, data=rq_data,
                                 verify=False)
    except Exception as e:
        print(e)
        return False
    if response.status_code != 201:
        print("[Fail] Can not get session token at url: {}. "
              "Code {}".format(url, response.status_code))
        return False
    print("[Success] Created session at url: {}".format(url))
    return response.headers


def dell_expire_session_token(url, headers):
    try:
        response = requests.delete(url, headers=headers, verify=False)
    except Exception as e:
        print(e)
        return False
    if response.status_code != 200:
        print("[Fail] Can not expire session at url: {}. "
              "Code {}".format(url, response.status_code))
    print("[Success] Expired session at url: {}".format(url))
    return None


def dell_get_basic_info(system_url, headers, redfish_version):
    basic_info = call_api_get(system_url, headers)
    if not basic_info:
        return False
    try:
        memory_status = basic_info['MemorySummary']['Status'].get('Health')
        if not memory_status:
            memory_status = basic_info['MemorySummary']['Status'].get(
                'HealthRollup')
        memory_gib = basic_info['MemorySummary']['TotalSystemMemoryGiB']
        if redfish_version < 160:
            memory_gib *= 1.074
        result = {
            'bios_version': basic_info.get('BiosVersion'),
            'model': basic_info.get('Model'),
            'memory_gib': int(memory_gib),
            'memory_status': memory_status,
            'processor': {
                'number': basic_info['ProcessorSummary']['Count'],
                'model': basic_info['ProcessorSummary']['Model']
            },
            'serial': basic_info.get('SKU'),
            'health': basic_info['Status']['Health']
        }
        return result
    except Exception as e:
        print("Can not get basic info because {}".format(e))
        return False


def dell_get_processor_info(base_url, system, headers):
    processor = call_api_get(base_url + system + "/Processors/", headers)
    if not processor:
        return False
    try:
        result, info, fail_part = {}, [], []
        for i in processor.get('Members'):
            cpu_url = base_url + i['@odata.id']
            cpu = call_api_get(cpu_url, headers)
            if not cpu:
                continue
            if cpu['Status']['Health'] != 'OK':
                fail_part.append({
                    'location': cpu.get('Socket'),
                    'health': cpu['Status']['Health'],
                    'state': cpu['Status']['State']
                })
            cpu_data = {
                'core': cpu.get('TotalCores'),
                'thread': cpu.get('TotalThreads'),
                'model': cpu.get('Model')
            }
            info.append(cpu_data)
        result.update({'fail_part': fail_part, 'info': info,
                       'number': len(processor.get('Members'))})
        return result
    except Exception as e:
        print("Can not get Processor info because {}".format(e))
        return False


def dell_get_fan_info(chassis_url, headers):
    thermal = call_api_get(chassis_url + "/Thermal/", headers)
    if not thermal:
        return False
    try:
        result, fail_part, number = {}, [], 0
        for i in thermal.get('Fans'):
            if i['Status'].get('State') != 'Absent':
                number += 1
                if i['Status'].get('Health') != 'OK':
                    fail_part.append({
                        'name': i.get('Name'),
                        'health': i['Status'].get('Health'),
                        'state': i['Status'].get('State')
                    })
        result.update({'number': number, 'fail_part': fail_part})
        return result
    except Exception as e:
        print("Can not get Fan info because {}".format(e))
        return False


def dell_get_memory_info(base_url, system, headers, redfish_version):
    mem = call_api_get(base_url + system + "/Memory/", headers)
    if not mem:
        return False
    result, info, fail_part = {}, [], []
    try:
        for i in mem.get('Members'):
            ram = call_api_get(base_url + i['@odata.id'], headers)
            if not ram:
                continue
            if ram.get('Status'):
                if ram['Status']['State'] == 'Absent':
                    continue
                if ram['Status']['Health'] != 'OK':
                    fail_part.append({
                        'location': ram.get('DeviceLocator'),
                        'health': ram['Status']['Health'],
                        'state': ram['Status']['State']
                    })
            capa = ram.get('CapacityMiB')
            if not capa:
                capa = ram.get('SizeMB')
            if redfish_version < 160:
                capa *= 1.074
                capa = int(capa)
            type = ram.get('MemoryDeviceType')
            if not type:
                type = ram.get('DIMMType')
            if not info:
                info.append({
                    'index': 0,
                    'capacity_mib': capa,
                    'type': type,
                    'count': 1
                })
            else:
                new = True
                for r in info:
                    count = r['count']
                    index = r['index']
                    if r['capacity_mib'] == capa:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                if new:
                    info.append({
                        'index': len(info),
                        'capacity_mib': capa,
                        'type': type,
                        'count': 1
                    })
        result.update({'fail_part': fail_part, 'info': info})
        return result
    except Exception as e:
        print("Can not get Memory info because {}".format(e))
        return False


def dell_get_logical_disk_info(base_url, volume_url, headers):
    all_ldisk = call_api_get(base_url + volume_url, headers)
    if not all_ldisk:
        return False
    result = {}
    result.update({'number': len(all_ldisk.get('Members'))})
    info, ldisk_data = [], {}
    try:
        for i in all_ldisk.get('Members'):
            ldisk_url = base_url + i['@odata.id']
            ldisk = call_api_get(ldisk_url, headers)
            if not ldisk:
                continue
            if ldisk.get('VolumeType') == 'RawDevice':
                continue
            if ldisk.get('VolumeType') == 'Mirrored':
                raid_type = 1
            else:
                raid_type = 0
            ldisk_data = {
                'raid': raid_type,
                'capacity_gb': int(round(ldisk['CapacityBytes'] / 1000000000)),
            }
            pdisk = ldisk.get('Links')
            if pdisk:
                pdisk = pdisk.get('Drives')
            else:
                ldisk_data.update({'disks': []})
                info.append(ldisk_data)
                continue
            disk = []
            ldisk_data.update({'physical_count': len(pdisk)})
            for d in pdisk:
                pdisk_info = call_api_get(base_url + d['@odata.id'], headers)
                if not pdisk_info:
                    continue
                disk_data = {
                    'capacity_gb': int(round(
                        pdisk_info['CapacityBytes'] / 1000000000)),
                    'type': pdisk_info['MediaType']
                }
                if disk_data['type'] == 'HDD':
                    disk_data.update({'speed': pdisk_info['RotationSpeedRPM']})
                else:
                    disk_data.update({'speed': None})
                if not disk:
                    disk.append(disk_data)
                for r in disk:
                    if disk_data['capacity_gb'] != r['capacity_gb'] \
                            or disk_data['type'] != r['type']:
                        disk.append(disk_data)
            ldisk_data.update({'disks': disk})
            info.append(ldisk_data)
        result.update({'info': info})
        return result
    except Exception as e:
        print("Can not get logical disk info because {}".format(e))
        return False


def dell_get_physical_disk_info(base_url, drives, headers):
    result = {}
    result.update({'number': len(drives)})
    info, fail_part = [], []
    try:
        for i in drives:
            pdisk = call_api_get(base_url + i['@odata.id'], headers)
            if not pdisk:
                continue
            if pdisk['Status']['Health'] != 'OK' \
                    or pdisk['Status']['State'] != 'Enabled':
                lmsg = "" + pdisk.get('Name')
                fail_part.append(strip(lmsg))
            disk_data = {
                'capacity_gb': int(round(pdisk['CapacityBytes'] / 1000000000)),
                'type': pdisk['MediaType']
            }
            if disk_data['type'] == 'HDD':
                disk_data.update({'speed': pdisk['RotationSpeedRPM']})
            else:
                disk_data.update({'speed': None})
            if not info:
                disk_data.update({'count': 1, 'index': 0})
                info.append(disk_data)
            else:
                new = True
                for d in info:
                    count = d['count']
                    index = d['index']
                    if d['speed'] \
                            and d['capacity_gb'] == disk_data['capacity_gb'] \
                            and d['type'] == disk_data['type'] \
                            and d['speed'] == disk_data['speed']:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                    elif not d['speed'] \
                            and d['capacity_gb'] == disk_data['capacity_gb'] \
                            and d['type'] == disk_data['type']:
                        count += 1
                        info[index].update({'count': count})
                        new = False
                        break
                if new:
                    disk_data.update({'count': 1, 'index': len(info)})
                    info.append(disk_data)

        result.update({'fail_part': fail_part, 'info': info})
        return result
    except Exception as e:
        print("Can not get physical disk info because {}".format(e))
        return False


def dell_get_disk_info(base_url, system, headers):
    raid_controller_info = call_api_get(base_url + system + "/Storage/",
                                        headers)
    if not raid_controller_info:
        return False
    storage_url = ""
    for i in raid_controller_info.get('Members'):
        url = i['@odata.id']
        if "RAID" in url:
            storage_url = url
    if not storage_url:
        print("Can not get raid controller info, fail to get Disk info")
        return False
    storage = call_api_get(base_url + storage_url, headers)
    if not storage:
        return False
    if storage.get('Volumes'):
        ldisk = dell_get_logical_disk_info(base_url,
                                           storage['Volumes'].get('@odata.id'),
                                           headers)
    else:
        ldisk = None
    if storage.get('Drives'):
        pdisk = dell_get_physical_disk_info(base_url, storage['Drives'],
                                            headers)
    else:
        pdisk = None
    return {'logical_disk': ldisk, 'physical_disk': pdisk}


def dell_get_power_info(base_url, chassis, headers):
    all_power = call_api_get(base_url + chassis + "/Power/", headers)
    if not all_power:
        return False
    result, info, fail_part = {}, [], []
    try:
        result.update({
            'number': len(all_power.get('PowerSupplies'))
        })
        for i in all_power.get('PowerSupplies'):
            if i['Status']['Health'] != 'OK' or \
                    i['Status']['State'] != 'Enabled':
                fp = {'name': i.get('Name')}
                fail_part.append(fp)
            capa = i.get('PowerCapacityWatts')
            if not capa:
                model = i.get('Model')
                if model:
                    capa = int(model.split(",")[1].replace("W", ""))
            power_data = {
                'capacity': capa,
            }
            info.append(power_data)
        redundancy = False
        if all_power.get('Redundancy'):
            redundancy = True
        result.update({'info': info, 'fail_part': fail_part,
                       'redundancy': redundancy})
        return result
    except Exception as e:
        print("Can not get Power info because {}".format(e))
        return False


def dell_get_network_adapter_info(base_url, system, headers):
    all_na = call_api_get(base_url + system + "/NetworkAdapters/", headers)
    if not all_na:
        return False
    result, info, fail_part = {}, [], []
    try:
        for i in all_na.get('Members'):
            na_url = base_url + i['@odata.id']
            na = call_api_get(na_url, headers)
            if not na:
                continue
            na_data = {
                'model': na.get('Model'), 'name': na.get('Name')
            }
            if not na.get('Status'):
                info.append(na_data)
                continue
            if na['Status'].get('Health') \
                    and na['Status'].get('Health') != 'OK':
                fail_part.append({
                    'name': na.get('Name'), 'state': na['Status'].get('State'),
                    'health': na['Status']['Health'], 'model': na.get('Model')
                })
            info.append(na_data)
        result.update({'number': len(all_na.get('Members')), 'info': info,
                       'fail_part': fail_part})
        return result
    except Exception as e:
        print("Can not get Network info because {}".format(e))
        return False


def dell_get_snmp_service_info(base_url, manager, headers):
    mn = call_api_get(base_url + manager + "/NetworkProtocol/", headers)
    if not mn:
        return False
    snmp = mn.get('SNMP')
    if not snmp:
        return False
    enable = snmp.get('ProtocolEnabled')
    if enable:
        result = {'state': 'Enabled', 'list_string': ''}
    else:
        result = {'state': 'Disabled', 'list_string': ''}
    return result


def dell_get_idrac_info(base_url, manager, headers):
    idrac = call_api_get(base_url + manager, headers)
    if not idrac:
        return False
    result = {'model': idrac.get('Model'),
              'version': idrac.get('FirmwareVersion')}
    return result


def dell_get_bios_config_info():
    return None


def dell_get_object_info(base_url, object, headers):
    object_dict = {
        'system': 'Systems',
        'chassis': 'Chassis',
        'manager': 'Managers'
    }
    oj = object_dict.get(object)
    if not oj:
        return False
    oj_info = call_api_get(base_url + "/redfish/v1/" + oj, headers)
    if not oj_info:
        return False
    oj_list = oj_info.get('Members')
    if not oj_list:
        return False
    oj_url = oj_list[0].get('@odata.id')
    return oj_url


def dell_get_all_info(ip_address, username, password):
    print("--> Getting information for Dell device {}".format(ip_address))
    base_url = "https://" + ip_address
    redfish_version = dell_get_redfish_version(base_url, username, password)
    if not redfish_version:
        return None
    header_info = dell_get_session_token(base_url, username, password,
                                         redfish_version)
    if not header_info:
        return None
    token = header_info.get('X-Auth-Token')
    headers = {"Content-Type": "application/json", "X-Auth-Token": token}
    try:
        system = dell_get_object_info(base_url, 'system', headers)
        chassis = dell_get_object_info(base_url, 'chassis', headers)
        manager = dell_get_object_info(base_url, 'manager', headers)
        result = {
            'base_info': dell_get_basic_info(base_url + system, headers,
                                             redfish_version),
            'network': dell_get_network_adapter_info(
                base_url, system, headers),
            'processor': dell_get_processor_info(base_url, system, headers),
            'fan': dell_get_fan_info(base_url + chassis, headers),
            'disk': dell_get_disk_info(base_url, system, headers),
            'power': dell_get_power_info(base_url, chassis, headers),
            'memory': dell_get_memory_info(base_url, system, headers,
                                           redfish_version),
            'snmp': dell_get_snmp_service_info(base_url, manager, headers),
            'firmware': dell_get_idrac_info(base_url, manager, headers),
            'bios_config': dell_get_bios_config_info()
        }
    except Exception as e:
        print("Can not get Dell device info cause {}".format(e))
        result = None
    dell_expire_session_token(base_url + header_info.get('Location'), headers)
    return result


def form_cpu_info(processor):
    if not processor:
        return None
    try:
        if processor.get('fail_part'):
            cpu = 'NOK. ' + str(processor['fail_part'])
        else:
            model = processor['info'][0]['model']
            for i in processor['info']:
                if i['model'] != model:
                    cpu = 'NOK. ' + str(processor['info'])
                    return cpu
            cpu = 'OK. ' + str(processor['number']) + 'x' + model
        return cpu
    except Exception as e:
        print("Can not form processor info because {}".format(e))
        return None


def form_fan_info(fan):
    if not fan:
        return None
    if fan.get('fail_part'):
        fan = 'NOK. ' + str(fan['fail_part'])
    else:
        fan = 'OK. ' + str(fan.get('number')) + ' FAN'
    return fan


def form_ram_info(memory, basic_info):
    if not memory or not basic_info:
        return None
    try:
        if memory.get('fail_part') \
                or basic_info.get('memory_status') != 'OK':
            mem = 'NOK. ' + str(memory['fail_part'])
        elif len(memory['info']) == 1:
            mem = 'OK. total: ' + str(basic_info['memory_gib']) + 'GB' \
                  + ', ' + str(memory['info'][0]['count']) + 'x' \
                  + str(int(memory['info'][0]['capacity_mib'] / 1024)) \
                  + 'GB ' + memory['info'][0]['type']
        else:
            mem = 'OK. total: ' + str(basic_info['memory_gib']) + 'GB' \
                  + ', ' + str(memory['info'])
        return mem
    except Exception as e:
        print("Can not form memory info cause {}".format(e))
        return None


def form_disk_info(all_disk_info):
    if not all_disk_info:
        return None
    disk_speed = {
        10000: ' 10K', 15000: ' 15K', 7200: ' 7K2', 5400: ' 5K4'
    }
    ldisk = all_disk_info.get('logical_disk')
    pdisk = all_disk_info.get('physical_disk')
    if not ldisk or not pdisk:
        return None
    try:
        if pdisk.get('fail_part'):
            result = 'NOK. ' + str(pdisk['fail_part'])
            return result
        have_raid = False
        ldisk_info = 'Logical: '
        for i in ldisk['info']:
            if i.get('raid'):
                have_raid = True
                if len(i['disks']) == 1:
                    disk_info = str(i['physical_count']) + 'x' \
                                + str(i['disks'][0]['capacity_gb']) + 'GB ' \
                                + i['disks'][0]['type']
                    if i['disks'][0].get('speed'):
                        disk_speed_str = disk_speed \
                            .get(i['disks'][0].get('speed'))
                        if disk_speed_str:
                            disk_info += disk_speed_str
                        else:
                            disk_info += str(i['disks'][0].get('speed'))
                else:
                    disk_info = str(i['disks'])
                disk_info = 'RAID ' + str(i['raid']) + ' - ' + disk_info
            else:
                disk_info = ''
            ldisk_info += disk_info + ', '
        pdisk_info, disk_info = 'Physical: ', ''
        for i in pdisk['info']:
            disk_info = str(i['count']) + 'x' + str(i['capacity_gb']) + 'GB ' \
                        + i['type']
            if i.get('speed'):
                disk_speed_str = disk_speed.get(i.get('speed'))
                if disk_speed_str:
                    disk_info += disk_speed_str
                else:
                    disk_info += str(i.get('speed'))
            if i['index'] != (len(pdisk['info']) - 1):
                disk_info += ', '
            pdisk_info += disk_info
        if have_raid:
            result = 'OK. ' + ldisk_info + pdisk_info
        else:
            result = 'OK. ' + pdisk_info
        return result
    except Exception as e:
        print("Can not form disk info cause {}".format(e))
        return None


def form_power_info(power_info):
    if not power_info:
        return None
    try:
        if power_info['fail_part']:
            result = 'NOK. ' + str(power_info['fail_part'])
        else:
            capacity = power_info['info'][0]['capacity']
            for i in power_info['info']:
                if i['capacity'] != capacity:
                    result = 'OK. ' + str(power_info['info'])
                    return result
            result = 'OK. ' + str(power_info['number']) + 'x' \
                     + str(capacity) + 'W'
        return result
    except Exception as e:
        print("Can not form power info cause {}".format(e))
        return None


def form_network_info(network):
    if not network:
        return None
    try:
        if network.get('fail_part'):
            result = 'NOK. ' + str(network['fail_part'])
        else:
            card = ""
            count = 1
            for c in network['info']:
                if c.get('model'):
                    if count < len(network['info']):
                        card += c['model'] + ', '
                    else:
                        card += c['model']
                count += 1
            result = 'OK. ' + str(network['number']) + ' card(s): ' \
                     + card
        return result
    except Exception as e:
        print("Can not form network adapter info cause {}".format(e))
        return None


def form_snmp_info(snmp):
    if not snmp:
        return None
    try:
        if snmp['state'] == 'Enabled':
            result = 'OK. State: ' + snmp['state'] + '. String: ' \
                     + str(snmp['list_string'])
        else:
            result = 'NOK. State: ' + snmp['state'] + '. String: ' \
                     + str(snmp['list_string'])
        return result
    except Exception as e:
        print("Can not form SNMP info cause {}".format(e))
        return None


def form_firmware_info(firmware, basic_info):
    if not firmware:
        return None
    result = firmware.get('model') + ": " + firmware.get('version') \
             + ". BIOS: " + basic_info.get('bios_version')
    return result


def form_bios_config_info(bios_config):
    if not bios_config:
        return None
    return str(bios_config)


def form_data(info):
    try:
        result = {}
        basic_info = info.get('base_info')
        if basic_info:
            result.update({
                'Serial': basic_info.get('serial'),
                'Model': basic_info.get('model'),
                'Health': basic_info.get('health')
            })
        cpu = form_cpu_info(info.get('processor'))
        fan = form_fan_info(info.get('fan'))
        mem = form_ram_info(info.get('memory'), basic_info)
        disk = form_disk_info(info.get('disk'))
        network = form_network_info(info.get('network'))
        power = form_power_info(info.get('power'))
        firm = form_firmware_info(info.get('firmware'), basic_info)
        snmp = form_snmp_info(info.get('snmp'))
        bios_config = form_bios_config_info(info.get('bios_config'))
        result.update({
            'CPU': cpu, 'Fan': fan, 'RAM': mem, 'Disk': disk,
            'Network Card': network, 'Firmware': firm,
            'BIOS_Config': bios_config, 'SNMP': snmp, 'Power': power
        })
        return result
    except Exception as e:
        print("Can not forming data cause {}".format(e))


def load_workbook(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(sheet_name)
    print("Number of device (row): {}".format(sheet.max_row - 1))
    list_server = []
    for row in sheet.iter_rows(min_row=2, max_col=7, max_row=sheet.max_row,
                               values_only=True):
        list_server.append({
            'ip_mm': row[1].strip(),
            'type_server': row[0],
            'username_mm': row[2].strip(),
            'password_mm': row[3].strip(),
            'ip_os': row[4].strip(),
            'pass_vtadmin': row[5].strip(),
            'pass_root': row[6].strip()
        })
    wb.close()
    return list_server


def save_workbook(server_data, file_path, os_server_data):
    fields = [
        'Result', 'IP_OS', 'Type', 'Health', 'Hostname', 'CPU', 'RAM', 'Disk',
        'Fan', 'Power',
        'Network Card', 'SNMP', 'BIOS_Config', 'Serial', 'Firmware',
        'Capacity', 'BIOS_Date', 'BIOS_Version', 'OS_Distribution',
        'Iptables_Status', 'Number_Iptables_rules',
        'Number_Iptables_rules_in_file', 'Bond',
        'Bond0', 'Bond1', 'IP_Manager', 'Product_Name', 'Vendor_Name',
        'Logical_Volume', 'HBA',
        'Define_in_dcim', 'Contract_in_dcim', 'Warranty_in_dcim',
        'License_in_dcim',
        'Verify_status_in_dcim', 'Monitored_in_dcim'
    ]
    wb = openpyxl.Workbook()
    sheet = wb.active
    row, col = 1, 1
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for f in fields:
        sheet.cell(row=row, column=col).value = f
        sheet.cell(row=row, column=col).fill = openpyxl.styles. \
            PatternFill(fgColor="00008000", fill_type='solid')
        sheet.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center")
        sheet.cell(row=row, column=col).border = thin_border
        col += 1

    row = 2
    for element in server_data:
        col = 1
        check = 'OK'
        for f in fields:
            value = element.get(f)
            if not value:
                try:
                    value = os_server_data[element['IP_OS']].get(f)
                except:
                    print("Cannot get about information OS server %s" % element[
                        'IP_OS'])
            sheet.cell(row=row, column=col).value = value
            if value and value.startswith('NOK'):
                check = 'NOK'
                sheet.cell(row=row, column=col).fill = openpyxl.styles. \
                    PatternFill(fgColor="FF0000", fill_type='solid')
            sheet.cell(row=row,
                       column=col).alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center")
            sheet.cell(row=row, column=col).border = thin_border
            col += 1
        sheet.cell(row=row, column=1).value = check
        if check != 'OK':
            sheet.cell(row=row, column=1).fill = openpyxl.styles. \
                PatternFill(fgColor="FF0000", fill_type='solid')
        row += 1

    wb.save(file_path)
    print("Done saving checklist. Check {}".format(file_path))


def load_workbook_os(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(sheet_name)
    print("Number of device (row): {}".format(sheet.max_row - 1))
    info_os_server = {}
    for row in sheet.iter_rows(min_row=2, max_col=17, max_row=sheet.max_row,
                               values_only=True):
        info = {row[1].strip(): {
            'Hostname': row[0].strip(),
            'Capacity': row[2].strip(),
            'BIOS_Date': row[3].strip(),
            'BIOS_Version': row[4].strip(),
            'OS_Distribution': row[5].strip(),
            'OS_Kernel': row[6].strip(),
            'Iptables_Status': row[7].strip(),
            'Number_Iptables_rules': row[8].strip(),
            'Number_Iptables_rules_in_file': row[9].strip(),
            'Bond0': row[10].strip(),
            'Bond1': row[11].strip(),
            'IP_Manager': row[12].strip(),
            'Product_Name': row[13].strip(),
            'Vendor_Name': row[14].strip(),
            'Logical_Volume': row[15].strip(),
            'HBA': row[16].strip()

        }}
        info_os_server.update(info)
    wb.close()
    return info_os_server


def get_info_os(file_ansible, file_inventory):
    loader = DataLoader()

    context.CLIARGS = ImmutableDict(tags={}, listtags=False, listtasks=False,
                                    listhosts=False,
                                    syntax=False, connection='ssh',
                                    module_path=None, forks=100,
                                    remote_user='vt_admin',
                                    private_key_file=None, ssh_common_args=None,
                                    ssh_extra_args=None, sftp_extra_args=None,
                                    scp_extra_args=None, become=True,
                                    become_method='sudo', become_user='root',
                                    verbosity=True,
                                    check=False, start_at_task=None)

    inventory = InventoryManager(loader=loader, sources=(file_inventory))

    variable_manager = VariableManager(loader=loader, inventory=inventory,
                                       version_info=CLI.version_info(
                                           gitinfo=False))

    pbex = PlaybookExecutor(playbooks=[file_ansible], inventory=inventory,
                            variable_manager=variable_manager, loader=loader,
                            passwords={})
    pbex.run()


def os_server_check(os_server, type_server):
    os_server['Bond'] = "NOK"
    if 'enabled' in str(os_server.get("Iptables_Status")):
        os_server['Iptables_Status'] = "OK. " + str(
            os_server.get("Iptables_Status"))
    else:
        os_server['Iptables_Status'] = "NOK." + str(
            os_server.get("Iptables_Status"))
    if type_server == 1:  # Compute
        os_server['Type'] = "Compute"
        print(os_server['HBA'])
        if os_server['HBA'].count("Offline") > 0:
            os_server['HBA'] = "NOK. " + os_server['HBA']
        else:
            os_server['HBA'] = "OK. " + os_server['HBA']
        if 'layer' in os_server['Bond0'] or 'layer' in os_server['Bond1']:
            os_server['Bond'] = "OK"
    if type_server == 2:  # Ceph
        os_server['Type'] = "Ceph"
        os_server['HBA'] = "OK. " + os_server['HBA']
        if 'layer' in os_server['Bond0'] and 'layer' in os_server['Bond1']:
            os_server['Bond'] = "OK"
    if type_server == 3:  # Bare-metal
        os_server['Type'] = "Bare-metal"
        if os_server['HBA'].count("Offline") > 0:
            os_server['HBA'] = "OK. " + os_server['HBA']
        else:
            os_server['HBA'] = "NOK. " + os_server['HBA']
        os_server['Bond'] = "OK"
    return os_server


def create_file_inventory(list_server_input):
    general = ['[all:vars]\n', 'ansible_ssh_port=22\n',
               'ansible_become_method=su\n', '[host_toolchecklist]\n']
    with open(CURRENT_DIR + "/inventory_toolchecklist", "w") as file:
        file.writelines(general)
        for server in list_server_input:
            ssh_server = server.get('ip_os') + ' ansible_password="' \
                         + server.get('pass_vtadmin') \
                         + '" ansible_become_password="' \
                         + server.get('pass_root') + '"\n'
            file.write(ssh_server)


def get_contract_dcim(baseurl, headers, device_id):
    try:
        params = {'device_id': device_id}
        contract = requests.get(baseurl + "/dcim/contracts/", headers=headers,
                                params=params).json()['results'][0]
        if "UNKNOW" in contract['product_id'] \
                or not contract['contract_number']:
            return "NOK. {'product_id': %s, 'contract_number': %s}" % \
                   (contract['product_id'], contract['contract_number'])
        else:
            return "OK. {'product_id': %s, 'contract_number': %s}" % \
                   (contract['product_id'], contract['contract_number'])
    except:
        msg = "NOK. Can't get info contract in DCIM"
        print(msg)
        return msg


def get_license_dcim(baseurl, headers, device_id):
    try:
        params = {'device_id': device_id}
        licenses = requests.get(baseurl + "/dcim/licenses/", headers=headers,
                                params=params).json()['results']
        msg_nok = ''
        for license in licenses:
            if "UNKNOW" in license['name'] or not license['name'] \
                    or not license['end_time']:
                msg_nok = "NOK. {'name': %s, 'start_time': %s, 'end_time': %s}"\
                          % (license['name'], license['start_time'],
                             license['end_time'])
            else:
                return "OK. {'name': %s, 'start_time': %s, 'end_time': %s}" % \
                       (license['name'], license['start_time'],
                        license['end_time'])
        return msg_nok
    except:
        msg = "NOK. Can't get info license in DCIM"
        print(msg)
        return msg


def get_warranty_dcim(baseurl, headers, device_id):
    try:
        params = {'device_id': device_id}
        warrantys = requests.get(baseurl + "/dcim/warranties/", headers=headers,
                                 params=params).json()['results']
        msg_nok = ''
        for warranty in warrantys:
            if 'UNKNOW' in warranty['start_date'] \
                    or not warranty['expiration_date']:
                msg_nok = "NOK. {'name': %s, 'start_date': %s, " \
                          "'expiration_date': %s}" \
                          % (warranty['name'], warranty['start_date'],
                             license['expiration_date'])
            else:
                return "OK. {'name': %s, 'start_date': %s, " \
                       "'expiration_date': %s}" \
                       % (warranty['name'], warranty['start_date'],
                          warranty['expiration_date'])
        return msg_nok
    except:
        msg = "NOK. Can't get info warranty in DCIM"
        print(msg)
        return msg


def get_info_dcim(ip_os, ip_mm, data, token):
    print("--> Getting information DCIM for {}".format(ip_os))
    baseurl = 'http://10.255.58.203/api'
    headers = {'Authorization': 'Token %s' % token}
    params_device = {'ip_address': ip_mm}
    params_instance = {'ip_address': ip_os}
    check_token = requests.get(baseurl, headers=headers).json()
    if not check_token.get('dcim'):
        print("Token %s invalid for call DCIM api" % token)
        data["Define_in_dcim"] = "NOK. Token invalid for call api DCIM"
        return data
    try:
        instance = requests.get(baseurl + "/dcim/instances/", headers=headers,
                                params=params_instance).json()['results'][0]
    except:
        print("Server %s not define in DCIM" % ip_os)
        data["Define_in_dcim"] = "NOK. Not define in DCIM"
        return data
    try:
        device = requests.get(baseurl + "/dcim/devices/", headers=headers,
                              params=params_device).json()['results'][0]
    except:
        print("Server %s not define in DCIM" % ip_mm)
        data["Define_in_dcim"] = "NOK. Not define in DCIM"
        return data
    data["Define_in_dcim"] = "OK"
    data["Contract_in_dcim"] = get_contract_dcim(baseurl, headers, device['id'])
    data['License_in_dcim'] = get_license_dcim(baseurl, headers, device['id'])
    data['Warranty_in_dcim'] = get_warranty_dcim(baseurl, headers, device['id'])
    data['Verify_status_in_dcim'] = "OK. Verified" \
        if instance['verify_status'].get('label') == 'Verified' \
        else "NOK. Chua verify"
    monitored = instance['monitored'].get("label")
    if monitored != "OFF":
        data['Monitored_in_dcim'] = "OK. %s" % monitored
    else:
        data['Monitored_in_dcim'] = "NOK. OFF"
    return data


if __name__ == "__main__":
    urllib3.disable_warnings()
    parser = argparse.ArgumentParser(
        description='Tool checklist device hardware')
    parser.add_argument(
        "-i", "--input", help='Excel input file. Default: input.xlsx',
        required=False, default='input.xlsx')
    parser.add_argument(
        "-s", "--sheet", help='Sheet of excel file to read. Default: Sheet1',
        required=False, default='input.xlsx')
    parser.add_argument(
        "-o", "--output", help='Excel output file. Default: output.xlsx',
        required=False, default='output.xlsx')
    parser.add_argument(
        "-t", "--dcim-token", help='Token get DCIM',
        required=False, default='930f83a2fd22214481b0ed7b0724ac1be3e1bfc0')
    args = vars(parser.parse_args())
    server_data = []
    list_server_input = load_workbook(args['input'], 'Sheet1')
    create_file_inventory(list_server_input)
    get_info_os(CURRENT_DIR + '/ansible_toolchecklist.yml',
                CURRENT_DIR + '/inventory_toolchecklist')
    os_server_data = load_workbook_os(
        CURRENT_DIR + "/ansible_toolchecklist.xlsx", 'Sheet')
    for sv in list_server_input:
        ip_os = sv.get('ip_os')
        ip_mm = os_server_data.get(ip_os).get('IP_Manager')
        user_mm = sv.get('username_mm')
        pw_mm = sv.get('password_mm')
        type_server = sv.get('type_server')
        os_server_data[ip_os] = os_server_check(os_server_data[ip_os],
                                                type_server)
        os_server_data[ip_os] = get_info_dcim(ip_os, ip_mm,
                                              os_server_data[ip_os],
                                              args['dcim_token'])
        sv_vendor = os_server_data[sv['ip_os']].get('Vendor_Name')
        if sv_vendor == 'HPE' or sv_vendor == 'HP':
            sv_info = hpe_get_all_info(ip_mm, user_mm, pw_mm)
        elif sv_vendor == 'FUJITSU':
            sv_info = fjs_get_all_info(ip_mm, user_mm, pw_mm)
        elif sv_vendor == 'Dell Inc.':
            sv_info = dell_get_all_info(ip_mm, user_mm, pw_mm)
        else:
            sv_info = None
        if sv_info:
            data = form_data(sv_info)
        else:
            data = {
                'Health': 'NOK. Could not get information for IP %s' % ip_mm}
        data.update({"IP_OS": ip_os})
        server_data.append(data)
    save_workbook(server_data, args['output'], os_server_data)
