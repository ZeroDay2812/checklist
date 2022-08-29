from ansible.module_utils.basic import AnsibleModule
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def main():
    module = AnsibleModule(
        argument_spec = dict(
            fields = dict(type = 'str',required = True),
            facts = dict(type = 'str',required = True),
            path = dict(type = 'str',default = '/tmp/ansible_toolchecklist.xlsx'),
        ),
        supports_check_mode = True,
    )
    raw = module.params['facts']
    fields = list(str.split(module.params['fields']))
    facts = list(raw.split("*"))
    path = module.params['path']
    data = []  
    for i in facts:
        if i != '' and i.isspace() ==  False:
            data.append(i)

    if os.path.exists(path):
        book = openpyxl.load_workbook(path)
        sheet_2 = book.active
        sheet_2.append(data)
        book.save(path)
    else:
        book = openpyxl.Workbook()
        sheet_2 = book.active
        sheet_2.append(fields)

        sheet_2.append(data)
        book.save(path)

    module.exit_json(changed=True)

if __name__ == '__main__':
    main()

